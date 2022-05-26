import os
import re
import sys
from openpyxl import Workbook, load_workbook
from openpyxl import workbook
import pandas as pd
import requests
import easygui
import datetime


class AutoGrader:

    def main(self):

        # ask for the csv file of submission that we will parse to get submission files
        self.location = easygui.fileopenbox('Select the csv file of submission')
        # self.location = r"C:\Users\syada\Downloads\Excel Revision Test(0122)-submissions.csv"

        # ask for the .sag file that we will parse to get our answer key
        self.keyFile = easygui.fileopenbox('Select the Rubric file for this Evaluation')
        # self.keyFile = "C:\AutoGrade Excel\\test2.txt"

        
        # ask for the max marks which can be given for the particular Evaluation
        self.score = easygui.integerbox('Enter the max marks for this Evaluation')
        # self.score = 50
        
        # ask for assignment folder containing the things to grade
        # self.Assignment = easygui.diropenbox('Select the Folder '
        #                                      'which contains the Students Assignments')
        # self.Assignment = "C:\AutoGrade Excel\\test4"
        self.Assignment = self.webscrap()
        self.sheetToGradeNum = 0
        self.sheetToGradeNumbers = []
        self.aSyntax = []
        self.readAssignmentKey(self.keyFile)
        self.pointLoss = 0
        for x in range(0, self.aSyntax.__len__()):
            for i in range(0, self.aSyntax[x].__len__()):
                print(self.aSyntax[x][i].__str__() + "\n")

        # gF = Graded File, the file we write the output to
        # self.gF = open(self.Assignment + 'Graded.txt', 'w')
        self.col1,self.col,self.col2, self.col3 = [], [], [], []

        # load in each paper and grade them!
        for file in os.listdir(self.Assignment):
            if file.endswith(".xlsx"):
                print(self.Assignment + '/' + file.__str__())
                try:

                    self.gradePaper(self.Assignment + '/' + file.__str__())
              

                    print(self.col1)
                    print(self.col2)
                    print(self.col3)
                    self.col.clear()

                except:
                    print(file.__str__() + ' had an error, OOPS:', sys.exc_info()[0])
                    # shutil.move(self.Assignment + '/' + file.__str__(), loc + '/' + file.__str__())
                    # os.remove(self.Assignment + '/' + file.__str__())
                    self.col2.append('Error occured in file')
                    self.col3.append(-1)
                    continue


        df = pd.DataFrame(list(zip(self.col1,[i for i in self.col2],self.col3)),columns=['Students','feedback',f'Scores(out of {self.score})'])
                # print(df)
        df.to_csv('C:\AutoGrade Excel\Output.csv',index=False)
    
    def webscrap(self):
        df = pd.read_csv(self.location)
        path = r'C:\AutoGrade Excel'
        csv_file = self.location.split('\\')[-1].split('.')[0]
        filename = os.path.join(path,csv_file)
        os.makedirs(filename, exist_ok=True)
        for i in range(len(df)) :
            r = requests.get(df.iloc[i, 4], allow_redirects=True)
            if os.path.exists(os.path.join(filename,df.iloc[i, 0])+'.xlsx'):
                print(f'file_{i+1} already there, skipping')
            else: 
                with open(os.path.join(filename,df.iloc[i, 0])+'.xlsx', 'wb') as f:
                    f.write(r.content)
                    print(f'downloaded {i+1}th file')
        return filename

    def gradePaper(self, path):

        # loading in the excel sheet
        fileName = path.replace(self.Assignment + '/', '')
        fileName = fileName.replace('.xlsx', '')
        # self.gF.write(fileName + '\n\n')
        self.col1.append(fileName)
        #       print("load wb")
        awb = load_workbook(path)
        #      print("load wb2")
        awb2 = load_workbook(path, data_only=True)
        #     print("done loading wb and wb2")
        asheets = []
        asheetsNotFormulas = []
        for s in awb._sheets:
            asheets.append(s)
        for s in awb2._sheets:
            asheetsNotFormulas.append(s)
        print('Start')

        # If we have at least one sheet, we drop info the grading loop
        if asheets.__len__() > 0:
            # self.gF.write('Number of Sheets: ' + asheets.__len__().__str__()+'\n')
            self.col.append('Number of Sheets: ' + asheets.__len__().__str__())
            # score = 60
            score = self.score
            # print("score:" + score.__str__())
            try:
                # for each sheet
                for sheetNum in range(0, self.sheetToGradeNumbers.__len__()):
                    self.sheetToGradeNum = self.sheetToGradeNumbers[sheetNum]
                    ws = asheets[self.sheetToGradeNum]
                    # self.gF.write(f'\nSheet:  {ws.title} --> Sheet Score: {round(score, 2).__str__()}\n')
                    self.col.append(f'Sheet:  {ws.title} --> Sheet Score: {round(score, 2).__str__()}')

                    ws2 = asheetsNotFormulas[self.sheetToGradeNum]
                    # Go through each question
                    for qNum in range(0, self.aSyntax[sheetNum].__len__()):
                        # print("Question Number: "+ qNum.__str__())
                        curQ = self.aSyntax[sheetNum][qNum]
                        #                        print("question: ",curQ)
                        # Check to see if it is a single-condition question, check count if so
                        if curQ.__len__() == 1:
                            # print("ayy")
                            if not self.checkStatement(curQ[0], ws, ws2, score, True):
                                score -= self.pointLoss
                                print(
                                    "Losing " + self.pointLoss.__str__() + " points, now at " + round(score,2).__str__() + " on question number " + qNum.__str__())
                                # print("false")

                            # print('done')

                        elif curQ.__len__() > 1:
                            # check each condition, if any fail then we break out of loop and subtract points
                            for i in range(0, curQ.__len__()):
                                #                    print("Current Q: " + curQ[i].__str__())
                                finalCondition = (i == curQ.__len__() - 1)
                                if not self.checkStatement(curQ[i], ws, ws2, score, finalCondition):
                                    # print('False')
                                    score -= self.pointLoss
                                    print(
                                        "Losing " + self.pointLoss.__str__() + " points, now at " + round(score,2).__str__() + " on question number " + qNum.__str__())
                                    break
                                else:
                                    continue
            #                                    print('True')

            except AttributeError as e:
                # self.gF.write('Student left cells blank\n\n\n')
                self.col.append('Student left cells blank')
                print('Student left cells blank: ' + str(e))

            print('~~~~~~~~~~~~~~')
            print('Score: ' + round(score,2).__str__())
            print('~~~~~~~~~~~~~~')
            # self.gF.write(f'\nFinal Score: ({round(score,2).__str__()}/{self.score})\n\n\n\n')
            self.col3.append(round(score,2))
            self.col2.append(self.col.copy())
            
            
    # Check for Statement stmt in worksheet ws
    def checkStatement(self, stmt, ws, ws2, score, finalCondition):
        # print("Checking Statement")
        # parse stmt into useful information
        cellToCheck = stmt[0]
        n = stmt[1]
        valToCheck = stmt[2]
        comment = stmt[3]
        pointVal = stmt[4]
        correctAnswer = stmt[5]

        self.pointLoss = pointVal
        # print("Comment:" +comment)

        # check the workbook for the desired statement count, fail and subtract score if is less
        if (n > 0):
            if valToCheck.upper() == "ANS":
                # self.gF.write(comment + '\n')
                self.col.append(comment)
                print(comment)
                # print("Statement less than desired")
                return False
            elif valToCheck.upper() != "XXX" and (str(ws[cellToCheck].value)).upper().count(valToCheck.upper()) < n:
                # print(ws[cellToCheck].value)
                # print(ws.title)
                # self.gF.write(comment + '\n')
                self.col.append(comment)
                print(comment)
                # print("Statement less than desired")
                return False
            elif finalCondition and correctAnswer != 'XX' and not self.isFloat(self.isDate(ws2[cellToCheck].value)) and str(
                    self.isDate(ws2[cellToCheck].value)).upper() != correctAnswer.upper():
                # self.gF.write(
                    # 'Answer did not match correct value in cell ' + cellToCheck + ' on sheet ' + ws.title.__str__() + ' but used the correct formulas (-'+ str(pointVal/2) +'pt)\n')
                self.col.append(f'Answer did not match correct value in cell {cellToCheck} on sheet {ws.title.__str__()} but used the correct formulas (-{str(pointVal/2)}pt)')
                print('Answer Wrong Not Decimal')
                # stmt[4]=1

                self.pointLoss = pointVal/2
                # offsetting the -5 from getting this wrong so its only -1... shuddup
                return False
            elif finalCondition and correctAnswer != 'XX' and self.isFloat(self.isDate(ws2[cellToCheck].value)):
                # if its a decimal we want to shave it off and check the first 6 decimal places
                ws2float = round(float(ws2[cellToCheck].value), 2)
                if self.isFloat(correctAnswer):
                    correctFloat = round(float(correctAnswer), 2)
                else:
                    correctFloat = correctAnswer
                # print(correctFloat.__str__()+" and ws2 is "+ws2float.__str__())
                if ws2float != correctFloat:
                    # self.gF.write(
                        # 'Answer did not match correct value in cell ' + cellToCheck + ' on sheet ' + ws.title.__str__() + ' but used the correct formulas (-'+ str(pointVal/2) +'pt)\n')
                    self.col.append(f'Answer did not match correct value in cell {cellToCheck} on sheet {ws.title.__str__()} but used the correct formulas (-{str(pointVal/2)}pt)')
                    # stmt[4]=1
                    # offsetting the -5 from getting this wrong so its only -1... shuddup
                    self.pointLoss = pointVal/2

                    print(correctFloat.__str__() + " and ws2 is " + ws2float.__str__())
                    # print('Answer Wrong')
                    return False
                else:
                    # print("Answer Correct")
                    return True
            else:
                # print("Answer Correct")
                return True
        else:
            if str(ws[cellToCheck].value).upper().count(valToCheck) >= -n:

                # self.gF.write(comment + '\n')
                self.col.append(comment)
                print(comment)
                # print("Statement less than desired")
                return False
            else:
                return True

    # Read in an assignment key to the self.aSyntax variable
    def readAssignmentKey(self, keyPath):
        aKey = open(keyPath, 'r')

        # Break down lines in the key we just opened
        lines = [line.rstrip('\n') for line in aKey]

        curSheet = 0
        state = 0

        newQuestion = []

        multConditions = False

        for i in range(0, lines.__len__()):

            # Get the current line we are parsing
            l = lines[i]

            # State 0 = Start new Sheet
            #############################################################
            if state == 0:
                if l[0] == '#':
                    # Skip Line, continue because this is a comment
                    continue
                elif l[0] == '=':
                    self.sheetToGradeNum = curSheet
                    self.sheetToGradeNumbers.append(int(l[1]))
                    # curSheet =0
                    self.aSyntax.append([])
                    state = 1

            # State 1 = Check for New Question
            #############################################################
            elif state == 1:
                if l[0] == '*':
                    # print('Start Multiple Condition Statement')
                    newQuestion = []
                    state = 3
                elif l[0] == '[':
                    newQuestion = []
                    state = 2
                elif l[0] == '#':
                    continue
                elif l[0] == "=":
                    curSheet = curSheet + 1
                    self.sheetToGradeNum = int(l[1])
                    self.sheetToGradeNumbers.append(int(l[1]))
                    # curSheet =0
                    self.aSyntax.append([])
                    state = 1
                else:
                    print('Error on line ' + i.__str__() + '), stuck in state 1')

            # State 2 = Read single statement in, fall back to S1
            #############################################################
            if state == 2:

                # parse into tokens separated by spaces
                parsedLine = re.split('_', l)

                newStatement = self.readStatement(parsedLine)
                newQuestion = [newStatement]
                self.aSyntax[curSheet].append(newQuestion)
                state = 1

            # State 3 = Read multiple statements in as a multi-condition, fall back to S1 when Question ends
            #############################################################
            elif state == 3:
                parsedLine = re.split('_', l)

                # Start of multi-condition
                if parsedLine[0].startswith('*['):
                    if not multConditions:
                        multConditions = True
                        parsedLine[0] = parsedLine[0].replace('*', '')
                        newQuestion.append(self.readStatement(parsedLine))

                    # if multConditions, we need to finish the last question and then start a new one
                    # as this is the start of a new multi-condition question
                    elif multConditions:
                        self.aSyntax[curSheet].append(newQuestion)
                        newQuestion = []
                        parsedLine[0] = parsedLine[0].replace('*', '')
                        newQuestion.append(self.readStatement(parsedLine))

                # continuation of multi-condition
                elif parsedLine[0].startswith('**[') and multConditions:
                    parsedLine[0] = parsedLine[0].replace('**', '')
                    newQuestion.append(self.readStatement(parsedLine))

                # fall out of multi-condition when the line starts with [ or *[
                elif (parsedLine[0].startswith('[') or parsedLine[0].startswith('*[')) and multConditions:
                    # print('End Multiple Condition Statement')
                    multConditions = False
                    self.aSyntax[curSheet].append(newQuestion)

                    # ensure that the line we are currently on gets read
                    newQuestion = [self.readStatement(parsedLine)]
                    self.aSyntax[curSheet].append(newQuestion)
                    state = 1

            # if a multi-condition is the last line of the assignment file
            if i == lines.__len__() - 1 and multConditions:
                self.aSyntax[curSheet].append(newQuestion)

    # Read a statement from the parsed line in readAssignmentKey
    def readStatement(self, parsedLine):

        # This is the structure of our statements
        # [CellToCheck, operator, ValueToCheckFor, Grader Comment, PointValue, Correct Answer]

        newStatement = ['',  # Cell To Check
                        0,  # Operator
                        '',  # Value we want to check is in Cell To Check
                        '',  # Grader's comment about getting this question condition wrong
                        0,  # The point value that this question condition is worth
                        '']  # The correct answer to compare against
        newStatement[0] = parsedLine[0][1:-1]
        # print(newStatement[0])

        ####Check Operator here####
        # if Check keyword is used, we search for soemthing that will result in a correct answer
        if 'Check' in parsedLine[1]:
            newStatement[1] = int(parsedLine[1][-1])
        # if Discard keyword is used, we search for something that will result in a wrong answer
        elif 'Discard' in parsedLine[1]:
            newStatement[1] = -1 * int(parsedLine[1][-1])
        else:
            print(parsedLine[1])

        # Value To Check For
        newStatement[2] = parsedLine[2][1:-1]

        # Point Value
        newStatement[4] = round(float(parsedLine[4][1:]),2)

        # the grader's comment
        commentString = parsedLine[3] + ' (-' + newStatement[4].__str__() + 'pts)'
        newStatement[3] = commentString

        # correct answer
        newStatement[5] = parsedLine[5]

        return newStatement

    def isDate(self,x):
        if isinstance(x,datetime.date):
            x= x.strftime('%d-%m-%Y')
            return x
        else:
            return x


    def isFloat(self, values):
        if str(values).upper() == "TRUE" or str(values).upper() == "FALSE":
            return False
        try:
            float(values)
            return True
        except ValueError:
            return False


AG = AutoGrader()

AG.main()
